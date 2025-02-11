from flask import Flask,render_template,request,make_response,redirect
import base64
from openpyxl import load_workbook,Workbook
SALT = 'tickettest'
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
from Crypto.Random import get_random_bytes
import json

KEY = 'TICKETTEST987654'
EVENT_NAME = 'TESTEVENT'
PATH = 'main.xlsx'

def aes_encrypt(plaintext, key=KEY):
    key = key.encode('utf-8')
    if len(key) not in [16, 24, 32]:
        raise ValueError("Key must be either 16, 24, or 32 bytes long")
    iv = get_random_bytes(AES.block_size)
    cipher = AES.new(key, AES.MODE_CBC, iv)
    padded_text = pad(plaintext.encode('utf-8'), AES.block_size)
    ciphertext = cipher.encrypt(padded_text)
    encrypted_text = base64.b64encode(iv + ciphertext).decode('utf-8')
    return encrypted_text

def aes_decrypt(encrypted_text, key=KEY):
    key = key.encode('utf-8')
    if len(key) not in [16, 24, 32]:
        raise ValueError("Key must be either 16, 24, or 32 bytes long")
    encrypted_data = base64.b64decode(encrypted_text)
    iv = encrypted_data[:AES.block_size]
    ciphertext = encrypted_data[AES.block_size:]
    cipher = AES.new(key, AES.MODE_CBC, iv)
    decrypted_text = unpad(cipher.decrypt(ciphertext), AES.block_size).decode('utf-8')
    return decrypted_text


def genTicketID(id:int,username:str,other_info:str=None):
    '''
    接收给定参数，生成票面id。
    参数：id/用户名/其他信息，其中其他信息缺省为None。
    '''
    StringOrigin = f"{EVENT_NAME}By-PyTicketManager {id} {username}"
    if other_info:
        StringOrigin = StringOrigin + f" {other_info}"
    return aes_encrypt(StringOrigin,KEY)

class FileOperator:

    def __init__(self,file_path):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            result = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):  # 跳过空行
                    continue
                row_data = {}
                for header, value in zip(headers, row):
                    if header == 'chksts':
                        row_data[header] = value == 1
                    else:
                        row_data[header] = value
                result.append(row_data)
            self.data = result
        except FileNotFoundError:
            print(f"Error: The file '{file_path}' does not exist.")
        except Exception as e:
            print(f"An error occurred while reading the file: {e}")

    def Sync(self, path=PATH):
        try:
            # 创建一个新的工作簿
            workbook = Workbook()
            sheet = workbook.active
            
            headers = list(self.data[0].keys())
            
            # 写入表头
            sheet.append(headers)
            
            # 写入数据行
            for item in self.data:
                row = [item[header] for header in headers]
                sheet.append(row)
            
            # 保存工作簿到指定路径
            workbook.save(path)
            print(f"Data successfully written to {path}")
        
        except Exception as e:
            print(f"An error occurred while writing to the file: {e}")

    def Remove(self,id:int):
        for item in self.data:
            if item['id']==id:
                self.data.remove(item)
                return
            
    def Search(self,username:str):
        for item in self.data:
            if item['usr']==username:
                return item['id']

    def List(self):
        return (self.data)

    def Add(self,username):
        num = self.data[-1]['id']+1
        tid = genTicketID(num,username)
        user = {'id': num, 'usr': username, 'tid': tid, 'chksts': False}
        self.data.append(user)
        return user
    
    def CheckTicket(self,id):
        for item in self.data:
            if item['id']==id:
                item['chksts']=True
    def ShowUserByID(self,id):
        for item in self.data:
            if item['id']==id:
                return item

fo = FileOperator(PATH)

app = Flask(__name__)

@app.route('/list',methods=['POST'])
def list_all():
    return json.dumps(fo.List())

@app.route("/")
def mainpage():
    return "Welcome to my API Program!"

@app.route("/add",methods=['POST'])
def add():
    val = request.form.get("username")
    fo.Add(val)
    fo.Sync()
    return str([{"return":0,"message":"Success!"}])

@app.route("/remove",methods=['POST'])
def remove():
    try:
        val = request.form.get("id")
        fo.Remove(int(val))
        fo.Sync()
        return str([{"return":0,"message":"Success!"}])
    except:
        return str([{"return":-1,"message":"Error!"}])

@app.route("/check",methods=['id'])
def check():
    try:
        val = request.form.get('id')
        fo.CheckTicket(val)
        fo.Sync()
        return str([{"return":0,"message":"Success!"}])
    except:
        return str([{"return":-1,"message":"Error!"}])
    
@app.route('/search/username/<usr>')
def search_usr(usr):
    return json.dumps(fo.ShowUserByID(fo.Search(usr)))

@app.route('/search/id/<int:id>')
def search_id(id):
    return json.dumps(fo.ShowUserByID(id))

if __name__ == '__main__':
    app.run(port=7788)